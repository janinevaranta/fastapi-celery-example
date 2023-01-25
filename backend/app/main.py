import time
import shutil
import uuid

from fastapi import FastAPI, UploadFile, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse
from starlette.background import BackgroundTask
from typing import Union
from pydantic import BaseModel

from celery.result import AsyncResult

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet import worksheet
from io import BytesIO
from pathlib import Path

from .utils import cleanup
from .worker import etask

app = FastAPI()

origins = [
  "http://frontend.localhost",
  "http://localhost"
]

# Basically ignore and skip CORS for
# painless frontend co-operation.
app.add_middleware(
  CORSMiddleware,
  allow_origins=origins,
  allow_methods=["*"],
  allow_headers=["*"]
)

@app.get("/")
async def index():
  return { "message": "Welcome to the API!" }


@app.get("/task/easy")
async def easy_task():
  # Simulate an easy difficulty task.
  result_of_easy_task = 10+10
  return { "message": "Done!" }


@app.get("/task/medium")
async def medium_task():
  # Simulate a medium difficulty task.
  time.sleep(5)
  return { "message": "Done!" }

@app.get("/task/hard")
async def hard_task():
  # Simulate a hard difficulty task.
  time.sleep(60)
  return { "message": "Done!" }

@app.post("/task/excel/uploadfile")
def excel_task(file: UploadFile): # NOTE: The path function must be a "def" as openpyxl doesnt support async operations. Otherwise this will BLOCK.
  """ An actual task to see how fastapi works with managing excel files.
  This route handles the file inside the FastAPI engine. No workers involved yet."""

  if file.content_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
    # We need to first convert the file to a readable form.
    f = file.file.read()
    xlsx = BytesIO(f)    
    # As a simple test open the excel file and write a new sheet to it.
    wb: Workbook = load_workbook(xlsx)
    ws: worksheet = wb.create_sheet("title")
    ws["A1"] = "Tadaa! It works!"

    # We need to save the file temporarely.
    random_name = str(uuid.uuid4()) # EXTREMELY UNLIKELY that two files have the same name.
    new_filename = random_name + ".xlsx"
    wb.save(filename=new_filename)
    # Send the file response that is downloadble.
    return FileResponse(
      new_filename, 
      filename="this_file_was_made_by_fastapi.xlsx",
      # Clean up the temporary file.
      background=BackgroundTask(cleanup, [new_filename])
    )
  else:
    raise HTTPException(status_code=400, detail="Wrong file type")

# Define a pydantic for the route below.
# We must define this to declare a response type to the route below.
class TaskResponse(BaseModel):
  id: str
  location: Union[Path, str]

@app.post("/task/excel/uploadfile_with_progress", status_code=202)
def excel_task_with_progress(file: UploadFile) -> TaskResponse: # Same thing here, no async. Otherwise this will block for other requests.
  """A task that utilizes workers. Once a file has been received in a request it is saved to a disk and
  information about it is sent to celery. Celery will then work on the file in the background."""

  if file.content_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
    # Save the file to a disk for the worker to read.
    random_name = str(uuid.uuid4()) # EXTREMELY UNLIKELY that two files have the same name.
    filepath = "/usr/src/app/uploads/"
    filename: str =  random_name + ".xlsx"

    full_filepath = filepath + filename

    with open(full_filepath, "wb+") as buffer:
      # We use shututil because it saves the file in chunks and is safer.
      shutil.copyfileobj(file.file, buffer)

    # Initialize the task. The work now moves to Celery server.
    task = etask.delay(filename)
    print("Task started:")
    print("ID: ", task.id)
    print("Filepath: ", full_filepath)

    # The client will receive the task id and "location" aka URL
    # where they can receive updates about the works current status.
    return {"id": task.id, "location": f"/queue/{task.id}"}
  else:
    raise HTTPException(status_code=400, detail="Wrong file type")

@app.get("/queue/{task_id}")
async def get_status(task_id):
  """URL used to receive updates on Celery tasks."""

  task_result = AsyncResult(task_id)
  result = {
    "id": task_id,
    "status": task_result.status,
    "result": task_result.result
  }
  print("ID: ", task_id)
  print("Status: ", task_result.status)
  print("Result: ", task_result.result) # Once the task is finished, this will point to the download URL.
  
  return JSONResponse(result)

@app.get("/downloads/{filename}")
async def file_downloads(filename):
  """Route for file downloads. The client should only come here after receiving a successfully 
  finished task."""

  filepath = "/usr/src/app/downloads/"
  full_filepath = Path(filepath + filename)

  if full_filepath.is_file():
    return FileResponse(
      full_filepath,
      filename="this_file_was_made_in_task.xlsx", # The name is only set in the header.
      background=BackgroundTask(cleanup, [full_filepath]) # Of course, we need to clean the place.
    )
  else:
    return HTTPException(404, "File not found")
  
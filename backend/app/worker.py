# Testing a worker queue workflow.
import os
from pathlib import Path
from celery import Celery, Task

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet import worksheet
from tempfile import NamedTemporaryFile

from .utils import cleanup

celery = Celery(__name__)
celery.conf.broker_url = os.environ.get("CELERY_BROKER_URL", "redis://localhost:6379")
celery.conf.result_backend = os.environ.get("CELERY_RESULT_BACKEND", "redis://localhost:6379")

@celery.task(name="excel_task")
def etask(filename: str):
  # Construct the actual filepath
  full_filepath = Path("/usr/src/app/uploads/" + filename)

  # As a simple test open the excel file and write a new sheet to it.
  wb: Workbook = load_workbook(full_filepath)
  ws: worksheet = wb.create_sheet("A new sheet")
  ws["A1"] = "Tadaa! It works!"

  # Save the file to a location on the disk.
  new_filepath = "/usr/src/app/downloads/"
  new_full_filepath = new_filepath + filename
  wb.save(new_full_filepath)

  # Cleanup the old file.
  cleanup([full_filepath])

  # Return the filename that will be passed to the client.
  return filename
  
  